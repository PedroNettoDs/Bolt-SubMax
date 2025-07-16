from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.db import OperationalError
from openpyxl import load_workbook
from Pages.models import Aluno, Plano, Avaliacao
from django.contrib import messages

# Lista alunos com busca
def alunos_view(request):
    """Exibe a lista de alunos com opção de busca por nome ou telefone."""
    search = request.GET.get('q', '')
    try:
        planos = Plano.objects.all()
    except OperationalError:
        planos = []
    try:
        if search:
            alunos = Aluno.objects.filter(nome__icontains=search) | Aluno.objects.filter(telefone__icontains=search)
        else:
            alunos = Aluno.objects.all()
    except OperationalError:
        alunos = []
    return render(request, 'paginas/alunos.html', {'alunos': alunos, 'planos': planos})

# Cadastra novo aluno
def cadastrar_aluno(request):
    """Cadastra um novo aluno com os dados fornecidos no formulário."""
    if request.method == 'POST':
        data_str = request.POST.get('data_nascimento')
        try:
            data_nascimento = datetime.strptime(data_str, '%Y-%m-%d').date() if data_str else None
        except (ValueError, OperationalError):
            data_nascimento = None

        plano = None
        plano_id = request.POST.get('plano')
        if plano_id:
            try:
                plano = Plano.objects.get(id=plano_id)
            except (Plano.DoesNotExist, OperationalError):
                plano = None

        aluno = Aluno(
            nome=request.POST.get('nome'),
            data_nascimento=data_nascimento,
            sexo=request.POST.get('sexo'),
            telefone=request.POST.get('telefone'),
            email=request.POST.get('email'),
            condicoes_medicas=request.POST.get('condicoes_medicas'),
            ciclo_menstrual=request.POST.get('ciclo_menstrual'),
            medicamentos=request.POST.get('medicamentos'),
            lesoes=request.POST.get('lesoes'),
            ocupacao=request.POST.get('ocupacao'),
            estilo_vida=request.POST.get('estilo_vida'),
            pratica_esportes=request.POST.get('pratica_esportes'),
            pratica_exercicios=request.POST.get('pratica_exercicios'),
            dias_disponiveis=request.POST.get('dias_disponiveis'),
            horario_preferencia=request.POST.get('horario_preferencia'),
            exercicio_preferencia=request.POST.get('exercicio_preferencia'),
            info_adicional=request.POST.get('info_adicional'),
            tempo_atual_treino=request.POST.get('tempo_atual_treino'),
            tempo_destreinado=request.POST.get('tempo_destreinado'),
            tempo_treino_anterior=request.POST.get('tempo_treino_anterior'),
            plano_tipo=plano
        )
        try:
            aluno.save()
            messages.success(request, "Aluno Cadastrada com sucesso!.")
        except OperationalError:
            pass
        return redirect('/alunos')

# Atualiza aluno
def atualizar_aluno(request, aluno_id):
    """Atualiza os dados de um aluno existente."""
    try:
        aluno = get_object_or_404(Aluno, id=aluno_id)
    except OperationalError:
        return redirect('/alunos')
    if request.method == "POST":
        aluno.nome = request.POST.get('nome')
        aluno.telefone = request.POST.get('telefone')
        aluno.sexo = request.POST.get('sexo')
        aluno.ativo = 'ativo' in request.POST
        aluno.email = request.POST.get('email')
        try:
            if request.POST.get('data_nascimento'):
                aluno.data_nascimento = datetime.strptime(request.POST['data_nascimento'], '%Y-%m-%d').date()
        except (ValueError, OperationalError):
            pass
        try:
            aluno.save()
            messages.success(request, "Usuario Editado com sucesso!")
        except OperationalError:
            pass
        return redirect('view_aluno', aluno_id=aluno.id)
    return render(request, 'aluno_atualizar.html', {'aluno': aluno})

# Função para listar alunos gerando tags (se houver campo 'etiquetas')
def lista_alunos(request):
    """Lista todos os alunos e gera tags para cada um, se o campo 'etiquetas' existir."""
    try:
        alunos = Aluno.objects.all()
    except OperationalError:
        alunos = []
    for aluno in alunos:
        aluno.tags = aluno.etiquetas.split(',') if hasattr(aluno, 'etiquetas') and aluno.etiquetas else []
    return render(request, 'paginas/alunos.html', {'alunos': alunos})

# Importa arquivo Excel e salva alunos no banco
def importar_excel(request):
    """Importa dados de alunos de um arquivo Excel e salva no banco de dados."""
    if request.method == 'POST' and request.FILES.get('arquivoExcel'):
        try:
            arquivo_excel = request.FILES['arquivoExcel']
            wb = load_workbook(arquivo_excel, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                nome = row[4].value
                data_nascimento = row[5].value
                sexo = row[6].value
                telefone = row[7].value
                email = row[8].value
                condicoes_medicas = row[9].value
                ciclo_menstrual = row[10].value
                medicamentos = row[11].value
                lesoes = row[12].value
                ocupacao = row[13].value
                estilo_vida = row[14].value
                pratica_esporte = row[15].value
                pratica_exercicio = row[16].value
                dias_disponiveis = row[17].value
                horario_preferido = row[18].value
                preferencia_exercicio = row[19].value
                informacao_adicional = row[20].value
                tempo_treinando = row[21].value
                tempo_destreinado = row[22].value
                tempo_treinamento_anterior = row[23].value

                if nome:
                    try:
                        Aluno.objects.create(
                            nome=nome or '',
                            data_nascimento=data_nascimento or None,
                            sexo=sexo or '',
                            telefone=telefone or '',
                            email=email or '',
                            condicoes_medicas=condicoes_medicas or '',
                            ciclo_menstrual=ciclo_menstrual or '',
                            medicamentos=medicamentos or '',
                            lesoes=lesoes or '',
                            ocupacao=ocupacao or '',
                            estilo_vida=estilo_vida or '',
                            pratica_esportes=pratica_esporte or '',
                            pratica_exercicios=pratica_exercicio or '',
                            dias_disponiveis=dias_disponiveis or '',
                            horario_preferencia=horario_preferido or '',
                            exercicio_preferencia=preferencia_exercicio or '',
                            info_adicional=informacao_adicional or '',
                            tempo_atual_treino=tempo_treinando or '',
                            tempo_destreinado=tempo_destreinado or '',
                            tempo_treino_anterior=tempo_treinamento_anterior or '',
                        )
                    except OperationalError:
                        pass
            return redirect('/alunos/')
        except OperationalError:
            return redirect('/alunos/')
    return redirect('/home/')

# TODO CORRIGIR: VALIDADOR DE ERRO

def atualizar_dados_complementares(request, aluno_id):
    """Atualiza os dados complementares de um aluno."""
    aluno = get_object_or_404(Aluno, id=aluno_id)

    if request.method == 'POST':
        aluno.condicoes_medicas = request.POST.get('condicoes_medicas')
        aluno.ciclo_menstrual = request.POST.get('ciclo_menstrual')
        aluno.medicamentos = request.POST.get('medicamentos')
        aluno.lesoes = request.POST.get('lesoes')
        aluno.ocupacao = request.POST.get('ocupacao')
        aluno.estilo_vida = request.POST.get('estilo_vida')
        aluno.pratica_esportes = request.POST.get('pratica_esportes')
        aluno.pratica_exercicios = request.POST.get('pratica_exercicios')
        aluno.dias_disponiveis = request.POST.get('dias_disponiveis')
        aluno.horario_preferencia = request.POST.get('horario_preferencia')
        aluno.exercicio_preferencia = request.POST.get('exercicio_preferencia')
        aluno.info_adicional = request.POST.get('info_adicional')
        aluno.tempo_atual_treino = request.POST.get('tempo_atual_treino')
        aluno.tempo_destreinado = request.POST.get('tempo_destreinado')
        aluno.tempo_treino_anterior = request.POST.get('tempo_treino_anterior')

        aluno.save()
        messages.success(request, "Dados pessoais atualizado com sucesso!")
        return redirect('view_aluno', aluno_id=aluno.id)

    # Se necessário, renderizar um template com erro ou redirecionar
    return redirect('view_aluno', aluno_id=aluno.id)
